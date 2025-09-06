import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_price_list_template():
    """Generate an Excel template for price list uploads."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Price List Template"
    
    # Define column headers
    headers = [
        "Product Name", 
        "Quality", 
        "Retail Price", 
        "Wholesale Price", 
        "Broker Price", 
        "Stock Quantity"
    ]
    
    # Define sample data
    sample_data = [
        ["Almonds", "Premium", 950, 850, 800, 100],
        ["Almonds", "Standard", 850, 750, 700, 150],
        ["Cashews", "Premium", 1200, 1100, 1050, 80],
        ["Pistachios", "Premium", 1500, 1400, 1350, 50],
        ["Walnuts", "Standard", 900, 800, 750, 120],
    ]
    
    # Add headers directly (no instruction row)
    header_row = 1
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}{header_row}'] = header
        ws[f'{col_letter}{header_row}'].font = Font(bold=True)
        ws[f'{col_letter}{header_row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws[f'{col_letter}{header_row}'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col_letter}{header_row}'].alignment = Alignment(horizontal='center')
    
    # Add sample data
    for row_num, row_data in enumerate(sample_data, header_row + 1):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Save to in-memory file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output
